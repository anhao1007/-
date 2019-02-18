# -*- coding:utf-8 -*- 
# @Time    :2019/1/1714:26
# @File    :.py
from openpyxl import Workbook;
from openpyxl import load_workbook;
from class_loanhttp.data_excel import Data_Excel
#读取表格
class Do_Excel:
    #初始化类的对象
    def __init__(self,file):
        self.file=file;

    #读取表的内容
    def read(self,table_name):
        #打开文件
        wk=load_workbook(self.file);
        #定位表单
        sheet=wk[table_name]
        do_case=[];#接收数据并且保存为列表类型
        for i in range(2,sheet.max_row+1):
            case=Data_Excel();
            case.case_id=sheet.cell(row=i, column=1).value #存的是case_ic
            case.title=sheet.cell(row=i, column=2).value#接受测试用例的标题
            case.url=sheet.cell(row=i, column=3).value#接收测试用例的http地址
            case.datas=sheet.cell(row=i, column=4).value#接收测试用例的数据
            case.methed=sheet.cell(row=i, column=5).value#接收http访问的方式
            case.expectations=sheet.cell(row=i, column=6).value#接收测试用例的期望值
            do_case.append(case)
        return do_case;
    #执行结果写入表格里面
    def write(self,table_name, row, actual,resutl):  # row行 col列 value_值
        wb = load_workbook(self.file);  # 打开文件
        sheet = wb[table_name];#定位表单
        sheet.cell(row, 7).value = actual;
        sheet.cell(row, 8).value = resutl;
        wb.save(self.file);  # 写入数据后保存





